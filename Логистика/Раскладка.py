import openpyxl


def main():
    input_file = input("Введите имя входного Excel-файла: ")
    output_file = input("Введите имя выходного Excel-файла: ")

    # Запрос параметров транспортного средства
    vehicle_length = float(input("Длина ТС (м): "))
    vehicle_width = float(input("Ширина ТС (м): "))
    max_weight_input = input("Грузоподъемность ТС (кг, 0 если без ограничений): ")
    vehicle_max_weight = float(max_weight_input) if max_weight_input != "0" else 0.0

    # Чтение данных из Excel
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    data = []

    headers = [cell.value for cell in sheet[1]]
    qty_col = headers.index("QTY") if "QTY" in headers else 0
    l_col = headers.index("L/m") if "L/m" in headers else 1
    w_col = headers.index("W/m") if "W/m" in headers else 2
    h_col = headers.index("H/m") if "H/m" in headers else 3
    weight_col = headers.index("GW (Kg)") if "GW (Kg)" in headers else 4

    for row in sheet.iter_rows(min_row=2, values_only=True):
        qty = row[qty_col]
        l = row[l_col] if isinstance(row[l_col], (int, float)) else 0.0
        w = row[w_col] if isinstance(row[w_col], (int, float)) else 0.0
        h = row[h_col] if isinstance(row[h_col], (int, float)) else 0.0
        weight = row[weight_col] if isinstance(row[weight_col], (int, float)) else 0.0

        # Обработка нескольких одинаковых грузов (QTY > 1)
        for _ in range(int(qty)):
            data.append({
                'l': float(l),
                'w': float(w),
                'weight': float(weight),
                'original_row': len(data) + 2  # Сохраняем исходный номер строки
            })

    # Сортировка грузов по убыванию максимального габарита
    sorted_data = sorted(data, key=lambda x: max(x['l'], x['w']), reverse=True)

    # Класс для управления полками в ТС
    class Shelf:
        __slots__ = ('remaining_length', 'height')

        def __init__(self, remaining_length, height):
            self.remaining_length = remaining_length
            self.height = height

    # Класс для управления транспортным средством
    class Vehicle:
        __slots__ = ('length', 'width', 'max_weight', 'shelves', 'current_y', 'current_weight')

        def __init__(self, length, width, max_weight):
            self.length = length
            self.width = width
            self.max_weight = max_weight
            self.shelves = []
            self.current_y = 0.0
            self.current_weight = 0.0

        def can_accept_weight(self, weight):
            return self.max_weight == 0 or self.current_weight + weight <= self.max_weight

        def try_place_item(self, l, w, weight):
            if not self.can_accept_weight(weight):
                return False

            for shelf in self.shelves:
                # Ориентация 1: без поворота
                if w <= shelf.height and l <= shelf.remaining_length:
                    shelf.remaining_length -= l
                    self.current_weight += weight
                    return True
                # Ориентация 2: с поворотом
                if l <= shelf.height and w <= shelf.remaining_length:
                    shelf.remaining_length -= w
                    self.current_weight += weight
                    return True
            return False

        def try_create_new_shelf(self, l, w, weight):
            if not self.can_accept_weight(weight):
                return False

            options = []
            # Вариант 1: без поворота (l - длина, w - высота полки)
            if l <= self.length and self.current_y + w <= self.width:
                options.append({'height': w, 'length_used': l})
            # Вариант 2: с поворотом (w - длина, l - высота полки)
            if w <= self.length and self.current_y + l <= self.width:
                options.append({'height': l, 'length_used': w})

            if not options:
                return False

            # Выбор варианта с минимальной высотой полки
            best_option = min(options, key=lambda x: x['height'])
            new_shelf = Shelf(
                remaining_length=self.length - best_option['length_used'],
                height=best_option['height']
            )
            self.shelves.append(new_shelf)
            self.current_y += best_option['height']
            self.current_weight += weight
            return True

    # Распределение грузов по ТС
    vehicles = []
    results = {}  # Словарь для результатов: original_row -> номер ТС

    for item in sorted_data:
        placed = False
        # Пытаемся разместить в существующих ТС
        for vehicle_idx, vehicle in enumerate(vehicles):
            if vehicle.try_place_item(item['l'], item['w'], item['weight']):
                results[item['original_row']] = vehicle_idx + 1
                placed = True
                break
            elif vehicle.try_create_new_shelf(item['l'], item['w'], item['weight']):
                results[item['original_row']] = vehicle_idx + 1
                placed = True
                break

        # Создаем новое ТС при необходимости
        if not placed:
            new_vehicle = Vehicle(vehicle_length, vehicle_width, vehicle_max_weight)
            vehicles.append(new_vehicle)
            if new_vehicle.try_create_new_shelf(item['l'], item['w'], item['weight']):
                results[item['original_row']] = len(vehicles)
            else:
                print(f"⚠️ Груз не поместился: строка {item['original_row']}")
                results[item['original_row']] = None

    # Запись результатов в Excel
    result_col = max(sheet.max_column, 6) + 1
    sheet.cell(row=1, column=result_col, value="Траспортное средство №")

    for row_idx in range(2, sheet.max_row + 1):
        if row_idx in results:
            sheet.cell(row=row_idx, column=result_col, value=results[row_idx])

    wb.save(output_file)
    print(f"✅ Результаты сохранены в файл: {output_file}")
    print(f"Использовано транспортных средств: {len(vehicles)}")


if __name__ == "__main__":
    main()